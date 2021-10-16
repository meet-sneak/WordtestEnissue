import json,requests,openpyxl,os
from openpyxl import load_workbook


user_data={"email":str(16521687458),"password":"admin@123"}

login_url='https://api.07future.com/api/auth'

class Loin():

    def __init__(self):
        pass

    def login(self,user_data=user_data):

        session = requests.session()
        resing = session.request('POST',login_url,json=user_data)
        print(resing.text)

        return session

class Handle_excel:

    def __init__(self,filename,sheetname = None):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):

        wb = load_workbook(filename = self.filename,read_only = True)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        sheetbook=[] #单表
        on_off=True
        for row in ws.rows: #行
            perlinecc = dict()
            if on_off:
                perline=[]
                for cell in row: #列
                    perline.append(cell.value)
                keyes=perline   #表头
                on_off=False
            else:
                i=0
                for cell in row: #列
                    perlinecc[keyes[i]]=cell.value
                    i += 1

                sheetbook.append(perlinecc)
        wb.close()
        return sheetbook    #返回结果示例：[{"key01":value01},{"key02":value02}]

    def write_data(self, data, sheetname=None): #data示例：[{"key01":value01},{"key02":value02}]

        wb = load_workbook(self.filename)
        if sheetname is None:
            ws = wb.active
        else:
            ws = wb[sheetname]
        on_off = True

        if on_off:
            k=1
            for j in data[0]:
                    ws.cell(1,k,value=j)    #表头
                    k+=1
            on_off=False

        for i in range(len(data)):
            k=1
            for j in data[i]:
                ws.cell(i+2,k,value=data[i][j])
                k+=1

        wb.save(self.filename)
        wb.close()

if __name__ == '__main__':

    session=Loin().login()
    school="yjd"
    pingce_url="https://api.07future.com/api/psychologicalTest/instant/teacher?school={0}&page=1&pageSize=2000".format(school)
    res=session.request("GET",pingce_url)
    res_date = json.loads(res.text)["instant"]

    total=json.loads(res.text)["count"]
    evaluating=[]

    for i in  range(total):

        details = dict(res_date[i]["classTotal"][0])
        details["学号"]=res_date[i]["student"]["sid"]
        details["姓名"] =res_date[i]["student"]["name"]
        details["测试时间"] =str(res_date[i]["lastUpdate"])[:10]
        details["结果"] =res_date[i]["reportAuto"]
        details["总分"] =res_date[i]["total"]
        details["测谎"] =res_date[i]["bad"]

        evaluating.append(details)

    print(json.dumps(evaluating,ensure_ascii=False))

    # filess="C:\\Users\\41470\\Documents\\雅加达心理评测.xlsx"
    # cc=Handle_excel(filess)
    # dd=cc.read_data()
    # print(json.dumps(dd,ensure_ascii=False))
    #
    # filessxx = "C:\\Users\\41470\\Documents\\雅加达心理评测01.xlsx"
    # ccxx=Handle_excel(filessxx)
    # ddxx = ccxx.write_data(dd)
    # print("运行完毕~")








