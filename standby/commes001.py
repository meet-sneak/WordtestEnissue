import requests,openpyxl,json,re
from bs4 import BeautifulSoup
from openpyxl import load_workbook

class Url_resources:


    def __init__(self):
        pass

    def url_resources(self,url,method='GET',headers=None,params=None,data=None):

        headers={
    'Content-Type':'application/json; charset=UTF-8',
    'Cookie':'JSESSIONID=c705c9e5de377425',
    'Host':'www.clutchprep.com',
    'Origin':'https://www.clutchprep.com',
    'Referer':'https://www.clutchprep.com/',
    'User-Agent':'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Mobile Safari/537.36',
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    }
        res = requests.request(method=method,url=url,headers=headers,params=params,data=data)
        soup = BeautifulSoup(res.text,'lxml')
        # ht_data = soup.select('.fp-engine')
        ht_data = soup.find_all(text=re.compile("videos/(.+?).mp4"))  #单次过滤的原始数据
        datacc=re.findall(r"videos/(.+?).mp4",str(ht_data))
        datacc=set(datacc)
        result=[]
        i=0
        for item in datacc:
            i=i+1
            if '_480p' or '_720p' not in item:
                sane_url='https://b-cdn-video.clutchprep.com/videos/'+item+'.mp4'
                result.append(sane_url)

        return ht_data,result,i

class Url_page:


    def __init__(self):
        pass

    def url_page(self,url,method='GET',headers=None,params=None,data=None):

        headers={
    'Content-Type':'application/json; charset=UTF-8',
    'Cookie':'JSESSIONID=c705c9e5de377425',
    'Host':'www.clutchprep.com',
    'Origin':'https://www.clutchprep.com',
    'Referer':'https://www.clutchprep.com/',
    'User-Agent':'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Mobile Safari/537.36',
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    }
        res = requests.request(method=method,url=url,headers=headers,params=params,data=data)
        soup = BeautifulSoup(res.text,'lxml')
        ht_data = soup.select('.single-subject__single-chapter__topic')
        datacc=re.findall(r"href=\"(.+?)\">",str(ht_data))
        datacc=set(datacc)
        result=[]
        i=0
        for item in datacc:
            i=i+1
            sane_url='https://www.clutchprep.com'+item
            result.append(sane_url)
            #print(sane_url) #每个页面路由

        return result,i

class Handle_excel:

    def __init__(self,filename,sheetname = None):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):

        wb = load_workbook(filename = self.filename,read_only = True)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        testcase_list=0
        cc=ws.cell(1,3)
        sheetbook=[] #单表
        for row in ws.rows: #行
            perline=[]
            for cell in row: #列
                perline.append(cell.value)
            sheetbook.append(perline)
        wb.close()
        return sheetbook

    def write_data(self,row, column, data, sheetname=None):

        wb = load_workbook(self.filename)
        if sheetname is None:
            ws = wb.active
        else:
            ws = wb[sheetname]
        ws.cell(row, column, value=data)
        wb.save(self.filename)
        wb.close()

class Down_binary:
    def __init__(self):
        pass

    def Down_binary(self,url,i='01',method='GET',path='./',type='mp4'):

        res = requests.request(method=method,url=url)
        content = res.content
        path=path+'cat0{}.{}'.format(i,type)
        with open(path,'wb') as f:
            f.write(content)

        return path

class Run:
    def __init__(self):
        pass

    def run(self,path,url=None):
        excel=Handle_excel(path,'Sheet1').read_data()
        all_url=[]
        for cc in excel:
            if cc[2] is not None:
                urles='https://www.clutchprep.com'+cc[2]
                all_url.append(urles)
        print('打印第一次',all_url) #表格中所有链接
        if len(all_url):
            if url:
                url='https://www.clutchprep.com/organic-chemistry'
                all_url,x=Url_page().url_page(url)
                print('打印第二次',all_url) #种子页面中所有链接

        video_url=[]
        for iurl in all_url:
            test01,html,i=Url_resources().url_resources(iurl)
            if len(html):
                video_url.append(html)
                print(html) #每个url的视频

        return  video_url

if __name__ == '__main__':
    video=Run().run(path='./chapter22.xlsx') #[参数,2个：(path：页面链接表格地址、类型为xlsx文件，url：目标网页地址)]
    pass



